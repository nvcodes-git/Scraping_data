from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pandas as pd
import time
import os

URL = "https://admision.unmsm.edu.pe/Website20262/A/A.html"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "output", "resultados_sanmarcos.xlsx")

options = webdriver.ChromeOptions()
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--headless")

driver = webdriver.Chrome(
    service=Service("/usr/bin/chromedriver"),
    options=options,
)

driver.get(URL)
time.sleep(3)

print("Page title:", driver.title)
print("Browser opened successfully.")

# --- Step 2: Collect all major links from the homepage ---

major_links = driver.find_elements(By.CSS_SELECTOR, "table a")
major_data = [(link.text.strip(), link.get_attribute("href")) for link in major_links]
print(f"Found {len(major_data)} majors.")

# --- Step 3: Visit each major and extract all applicants ---

all_applicants = []

for major_name, major_url in major_data:
    print(f"Scraping: {major_name} ...", end=" ", flush=True)
    driver.get(major_url)
    time.sleep(3)

    table_id = driver.execute_script("""
        var tables = $.fn.dataTable.tables();
        return tables.length > 0 ? tables[0].id : null;
    """)

    if not table_id:
        print("No DataTable found, skipping.")
        continue

    driver.execute_script(f"$('#{table_id}').DataTable().page.len(-1).draw();")
    time.sleep(3)

    rows = driver.find_elements(By.CSS_SELECTOR, f"#{table_id} tbody tr")
    count = 0
    for row in rows:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) >= 6:
            all_applicants.append({
                "Código": cols[0].text.strip(),
                "Apellidos y Nombres": cols[1].text.strip(),
                "Escuela": cols[2].text.strip(),
                "Puntaje": cols[3].text.strip(),
                "Mérito E.P": cols[4].text.strip(),
                "Observación": cols[5].text.strip(),
            })
            count += 1
    print(f"{count} applicants.")

driver.quit()

print(f"\nTotal applicants across all majors: {len(all_applicants)}")

# --- Step 4: Export to Excel ---

df = pd.DataFrame(all_applicants)
df["Puntaje"] = pd.to_numeric(df["Puntaje"], errors="coerce")
df["Mérito E.P"] = pd.to_numeric(df["Mérito E.P"], errors="coerce")

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Resultados")

    ws = writer.sheets["Resultados"]

    # Column widths
    col_widths = {"A": 12, "B": 40, "C": 40, "D": 12, "E": 12, "F": 35}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    passed_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        observacion = row[5].value or ""
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if "VACANTE" in observacion:
                cell.fill = passed_fill

    # Freeze header row
    ws.freeze_panes = "A2"

print(f"Excel saved to: {OUTPUT_PATH}")
